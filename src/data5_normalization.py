# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import shutil

import sys
sys.path.insert(0, '/normalization_function')

temp_dir = './temp/'
if os.path.exists(temp_dir):
    shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)
    print('Clear the directory: {}\n'.format(temp_dir))
else:
    os.makedirs(temp_dir)
    print('Create the directory: {}\n'.format(temp_dir))

# open file
loadfile = '../data/data5.xlsx'
loadbook = load_workbook(loadfile)
# get all sheets in the file
sheets = loadbook.get_sheet_names()
# get the first sheet name
sheet0 = sheets[0]

# create new file
writefile = './temp/data5_normalization.xlsx'
writebook = Workbook()

# get sheets
loadsheet = loadbook.get_sheet_by_name(name = sheet0)
writesheet = writebook.active

# normalize

starttimenormalize = time_normalization_datestr(loadsheet, 3, 1, 1)
owner_dictionary, ownernormalize = normalization(loadsheet, 3, 2, 2)
type_dictionary, typenormalize = normalization(loadsheet, 3, 4, 4)
reason_dictionary, reasonnormalize = normalization(loadsheet, 3, 5, 5)
endtimenormalize = time_normalization_datestr(loadsheet, 3, 6, 6)
status_dictionary, statusnormalize = normalization(loadsheet, 3, 7, 7)

print(owner_dictionary)

# write into sheet
write(writesheet, starttimenormalize, 1)
write(writesheet, ownernormalize, 2)
write(writesheet, typenormalize, 3)
write(writesheet, reasonnormalize, 4)
write(writesheet, endtimenormalize, 5)
write(writesheet, statusnormalize, 6)

# save file
writebook.save(writefile)

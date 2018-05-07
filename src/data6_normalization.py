# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import shutil

import sys
sys.path.insert(0, '/normalization_function')

temp_dir = './temp/'
if os.path.exists(temp_dir):
    shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)
    print('Clear the directory: {}\n'.format(temp_dir))
else:
    os.makedirs(temp_dir)
    print('Create the directory: {}\n'.format(temp_dir))

# open file
loadfile = '../data/data6.xlsx'
loadbook = load_workbook(loadfile)
# get all sheets in the file
sheets = loadbook.get_sheet_names()
# get the first sheet name
sheet0 = sheets[0]

# create new file
writefile = './temp/data6_normalization.xlsx'
writebook = Workbook()

# get sheets
loadsheet = loadbook.get_sheet_by_name(name = sheet0)
writesheet = writebook.active

# normalize
normalization_dictionary, normalize = normalization(loadsheet, 2, 4, 4)
starttimenormalize = time_normalization(loadsheet, 2, 5, 5)
endtimenormalize = time_normalization(loadsheet, 2, 6, 6)

print(normalization_dictionary)

# write into sheet
write(writesheet, normalize, 1)
write(writesheet, starttimenormalize, 2)
write(writesheet, endtimenormalize, 3)

# save file
writebook.save(writefile)
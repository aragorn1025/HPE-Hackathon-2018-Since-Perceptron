# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import shutil

import sys
sys.path.insert(0, '/normalization_function')

temp_dir = './temp/'
if os.path.exists(temp_dir):
    shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)
    print('Clear the directory: {}\n'.format(temp_dir))
else:
    os.makedirs(temp_dir)
    print('Create the directory: {}\n'.format(temp_dir))

# open file
loadfile = '../data/data3.xlsx'
loadbook = load_workbook(loadfile)
# get all sheets in the file
sheets = loadbook.get_sheet_names()
# get the first sheet name
sheet0 = sheets[0]

# create new file
writefile = './temp/data3_normalization.xlsx'
writebook = Workbook()

# get sheets
loadsheet = loadbook.get_sheet_by_name(name = sheet0)
writesheet = writebook.active

# normalize

mathod_dictionary, mathodnormalize = normalization(loadsheet, 4, 13, 13)
type_dictionary, typenormalize = normalization(loadsheet, 4, 14, 14)
height_dictionary, heightnormalize = normalization(loadsheet, 4, 15, 15)
timenormalize = time_normalization_datetime(loadsheet, 4, 16, 16)

print(owner_dictionary)

# write into sheet
write(writesheet, mathodnormalize, 1)
write(writesheet, typenormalize, 2)
write(writesheet, heightnormalize, 3)
write(writesheet, timenormalize, 4)

# save file
writebook.save(writefile)



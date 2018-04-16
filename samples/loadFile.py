# -*- coding: utf-8 -*-
"""
Created on Tue Apr 10 21:42:32 2018

@author: KoI
"""

from openpyxl import load_workbook

# open file
wb = load_workbook(filename = 'filename')
# get all sheets in the file
sheets = wb.get_sheet_names()
# get the first sheet name
sheet0 = sheets[0]

# get the first sheet
ws = wb.get_sheet_by_name(name = sheet0)
# get the row number from the sheet
rows = ws.rows
# get the column number
columns = ws.columns

content = []

for row in rows:
    # get each row's column value
    line = [col.value for col in row]
    # add into the list
    content.append(line)
    
print(content)

#for row in ws.iter_rows():
#    for cell in row:
#        print(cell.value, end='\t')
#    print(end='\n')

# print (ws.cell(row=2, column=1).value)

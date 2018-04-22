# -*- coding: utf-8 -*-
"""
Created on Sun Apr 22 23:32:57 2018

@author: KoI
"""

from openpyxl import load_workbook

# open file
file = 'data/data01.xlsx'
wb = load_workbook(file)
# get all sheets in the file
sheets = wb.get_sheet_names()
# get the first sheet name
sheet0 = sheets[0]

# create new sheet
if len(sheets) <2:
    wb.create_sheet('Normalization', index=1)
    sheets = wb.get_sheet_names()

sheet1 = sheets[1]

# get sheets
ws = wb.get_sheet_by_name(name = sheet0)
writesheet = wb.get_sheet_by_name(name = sheet1)
# get the row number from the sheet
rows = ws.rows
# get the column number
columns = ws.columns

content = []

# get 故障原因
for col in ws.iter_cols(min_row=2, max_col=4, min_col=4):
    for cell in col:
        content.append(cell.value)
#    print(content)

dic = []
match = False

# get all value in content
for i in content:
    for j in dic:
        if i == j:
            match = True
            break
    if match == False:
        dic.append(i)
    match = False

# normalize
normalize = []
for i in content:
    for j in dic:
        if i == j:
            normalize.append(dic.index(j))

# normalizec = list(zip(normalize))
print(normalize)

# make dictionary
d = {}
count = 0
for i in dic:
    d[count] = i
    count = count + 1
print(d)

# write in sheet
writesheet.append(normalize)
# save file
wb.save(file)